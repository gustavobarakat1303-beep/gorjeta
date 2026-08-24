#!/usr/bin/env python3
"""
Conciliacao mensal Pe de Manga.

Hierarquia dos controles:
  Excel (Planilha manual)     -> controle PRIMARIO (source of truth)
  Everest (export)            -> controle SECUNDARIO (deveria bater com o Excel)
  Fatura Itau                 -> complemento (tudo deveria ja estar nos 2 acima)

Estrategia de matching (3 rounds por cruzamento):
  1. Match EXATO por valor (multiset)
  2. Match APROXIMADO: mesmo fornecedor (normalizado + alias + fuzzy) + valor com
     tolerancia (max 5% ou R$ 50) + data proxima (<= 10 dias)
  3. Sobra: itens que nao casaram em nenhum round

Uso:
  python3 reconcile.py \
    --excel  caminho/excel_controle.xlsx \
    --everest caminho/everest_notas.xlsx \
    --fatura caminho/fatura_itau.xlsx \
    --out caminho/saida_conciliacao.xlsx \
    --mes 07/2026

--excel eh obrigatorio; --everest e --fatura opcionais mas normalmente ambos entram.
"""
import argparse
import json
import re
import sys
import os
from collections import defaultdict, Counter
from datetime import datetime, date
from difflib import SequenceMatcher

# openpyxl e importado onde eh usado (para permitir --help sem ele instalado)

# =============================================================================
# CONFIG e CONSTANTES
# =============================================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(os.path.dirname(SCRIPT_DIR), 'data')
ALIASES_FILE = os.path.join(DATA_DIR, 'supplier_aliases.json')

# Impostos: excluir do Everest quando o usuario disser "sem impostos"
TAX_RE = re.compile(r'\b(SIMPLES NACIONAL|GARE|ICMS|FGTS|INSS|IRRF|IRPJ|PIS|COFINS|'
                    r'DARF|DAS|ISS|IPTU|IPVA|IMPOSTO|TRIBUTO|SEFAZ|GNRE|CSLL|ISSQN|CPP|GPS)\b')

# Palavras que sinalizam servico/assinatura (sem DANFE esperado) na FATURA
SERVICO_FATURA = ['GOOGLE','FACEBK','FACEBOOK','ADOBE','CANVA','SPOTIFY','OPENAI',
                  'ANTHROPIC','CLAUDE','RYZE','CHATGPT','TRIPADVISOR','INDEED',
                  'TIM ','TIM*','JIM.COM','NUV ','SETTI','DM ','EBN ','MERCADOLIVRE',
                  'MERCADO*','LALAMOVE','HIGGSFIELD','ALIEXPRESS','MAGALU','MAGAZINE',
                  'UBER','99*','TIM']
COMB_ESTAC = ['POSTO','STONIC','ESTACIONAM','R.M.C.']

# Setores do Excel que sao servicos (sem DANFE esperado)
SERV_SETOR = {'MARKETING','EVENTUAL','MANUTENÇÃO','FUNCIONÁRIO EXTRA'}

# Palavras no fornecedor do Excel que indicam servico
SERV_NOME_EXCEL = ['GOOGLE','OPENAI','ADOBE','SPOTIFY','CANVA','LINKEDIN','TRIPADVISOR',
                   'INSTAGRAM','FACEBOOK','DESPESAS DIVERSAS','SWITCH APP','TIM',
                   'MERCADO LIVRE','MERCADOLIVRE','DELLAS','ROSOCHASNKY','ANTHROPIC',
                   'CLAUDE','RYZE','CHATGPT','HIGGSFIELD','MAGALU','MAGAZINE','UBER','99']

# =============================================================================
# NORMALIZACAO E FUZZY MATCH DE FORNECEDORES
# =============================================================================
_SUFFIX_RE = re.compile(r'\b(LTDA(?:\.| ME| EPP)?|S/?A|S\.A\.?|SA|EIRELI|ME|EPP|EPP\.?|-\s*FIXO)\b')
_PUNCT_RE = re.compile(r'[.,\-\/\*\(\)\[\]\|:;\'\"]+')
_SPACE_RE = re.compile(r'\s+')

def normalize_supplier(name):
    """Normaliza nome: uppercase, remove sufixos LTDA/SA/etc, remove pontuacao,
    colapsa espacos."""
    if not name: return ''
    s = str(name).upper().strip()
    s = _PUNCT_RE.sub(' ', s)
    s = _SUFFIX_RE.sub(' ', s)
    s = _SPACE_RE.sub(' ', s).strip()
    return s

def load_aliases():
    """Carrega o dicionario de aliases e monta um mapa reverso variante->canonico."""
    if not os.path.exists(ALIASES_FILE):
        return {}
    try:
        with open(ALIASES_FILE, encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f'WARN: nao consegui ler {ALIASES_FILE}: {e}', file=sys.stderr)
        return {}
    reverse = {}
    for canonical, variants in data.items():
        if canonical.startswith('_'): continue
        for v in variants:
            reverse[normalize_supplier(v)] = canonical
    return reverse

def canonicalize(name, aliases_map):
    """Retorna o nome canonico se existir alias, senao retorna o normalizado."""
    n = normalize_supplier(name)
    return aliases_map.get(n, n)

def supplier_similarity(a, b, aliases_map):
    """Retorna score 0..1 de similaridade entre dois nomes de fornecedor.
    1.0 se sao o mesmo canonico (via alias); senao usa Jaccard de tokens + SequenceMatcher."""
    if not a or not b: return 0.0
    ca = canonicalize(a, aliases_map)
    cb = canonicalize(b, aliases_map)
    if ca == cb and ca:
        return 1.0
    na = normalize_supplier(a)
    nb = normalize_supplier(b)
    if not na or not nb: return 0.0
    # Jaccard de tokens (min 3 chars)
    ta = {t for t in na.split() if len(t) >= 3}
    tb = {t for t in nb.split() if len(t) >= 3}
    if ta and tb:
        jac = len(ta & tb) / len(ta | tb)
    else:
        jac = 0.0
    # SequenceMatcher
    sm = SequenceMatcher(None, na, nb).ratio()
    return max(jac, sm)

# =============================================================================
# CARREGADORES
# =============================================================================
def load_excel_controle(path):
    """Carrega Excel controle mensal. Detecta header 'Fornecedor Data Valor Setor'
    em qualquer linha. Retorna lista de dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]  # sempre a primeira aba
    # achar linha do header
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [str(c).upper().strip() if c else '' for c in row]
        if 'FORNECEDOR' in vals and 'VALOR' in vals and ('DATA' in vals or 'DT' in vals):
            header_row = i
            break
    if header_row is None:
        raise ValueError(f'Nao achei header (Fornecedor/Data/Valor/Setor) no Excel {path}')
    # mapear indices
    headers = [str(c).upper().strip() if c else '' for c in list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]]
    idx = {name: headers.index(name) for name in ['FORNECEDOR','DATA','VALOR','SETOR'] if name in headers}
    if 'SETOR' not in idx:
        idx['SETOR'] = None
    rows = []
    for r in ws.iter_rows(min_row=header_row+1, values_only=True):
        forn = r[idx['FORNECEDOR']] if idx.get('FORNECEDOR') is not None else None
        valor = r[idx['VALOR']] if idx.get('VALOR') is not None else None
        data = r[idx['DATA']] if idx.get('DATA') is not None else None
        setor = r[idx['SETOR']] if idx.get('SETOR') is not None else None
        # filtrar linha de total (sem fornecedor) e vazias
        if forn is None: continue
        if valor is None: continue
        try:
            v = round(float(valor), 2)
        except (TypeError, ValueError):
            continue
        rows.append(dict(fornecedor=str(forn).strip(), data=data, valor=v, setor=setor))
    return rows

def load_everest_notas(path, exclude_taxes=False):
    """Carrega 'Manutencao de Notas Recebidas' OU 'Carteira de Titulos a Pagar' do Everest.
    Detecta o formato pelo cabecalho. Retorna lista de dicts uniforme.
    Se exclude_taxes=True, remove entradas de impostos."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    # ler primeira linha (header)
    hdr = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    hmap = {str(c).upper().strip(): i for i, c in enumerate(hdr) if c}
    rows = []
    tipo_export = None

    # tenta detectar 'Manutencao de Notas' (com Fantasia Fornecedor + V. Total)
    if 'FANTASIA FORNECEDOR' in hmap and 'V. TOTAL' in hmap:
        tipo_export = 'Manutencao'
        for r in ws.iter_rows(min_row=2, values_only=True):
            forn = r[hmap['FANTASIA FORNECEDOR']]
            if not forn: continue
            vtot = r[hmap['V. TOTAL']]
            if vtot is None: continue
            rows.append(dict(
                fornecedor=str(forn).strip(),
                tipo=r[hmap.get('TIPO FORNECEDOR')] if 'TIPO FORNECEDOR' in hmap else None,
                numero=r[hmap.get('NÚMERO')] if 'NÚMERO' in hmap else (r[hmap.get('NUMERO')] if 'NUMERO' in hmap else None),
                danfe=r[hmap.get('DANFE')] if 'DANFE' in hmap else None,
                emissao=r[hmap.get('D. EMISSÃO')] if 'D. EMISSÃO' in hmap else (r[hmap.get('D. EMISSAO')] if 'D. EMISSAO' in hmap else None),
                valor=round(float(vtot), 2),
                lancamento=r[hmap.get('D. LANÇAMENTO')] if 'D. LANÇAMENTO' in hmap else (r[hmap.get('D. LANCAMENTO')] if 'D. LANCAMENTO' in hmap else None),
                portador=None,
                origem=None,
                imposto=False,
            ))
    # tenta 'Carteira de Titulos a Pagar' (Razao Fornecedor + V. Original)
    elif 'RAZÃO FORNECEDOR' in hmap or 'RAZAO FORNECEDOR' in hmap:
        tipo_export = 'Carteira'
        forn_key = 'RAZÃO FORNECEDOR' if 'RAZÃO FORNECEDOR' in hmap else 'RAZAO FORNECEDOR'
        valor_key = None
        for k in ('V. ORIGINAL','V. SALDO','V. ATUALIZADO','V. TOTAL','VALOR'):
            if k in hmap: valor_key = k; break
        if not valor_key:
            raise ValueError('Carteira sem coluna de valor. Reexporte incluindo V. Original / V. Total.')
        for r in ws.iter_rows(min_row=2, values_only=True):
            forn = r[hmap[forn_key]]
            if not forn: continue
            vtot = r[hmap[valor_key]]
            if vtot is None: continue
            is_tax = bool(TAX_RE.search(str(forn).upper()))
            rows.append(dict(
                fornecedor=str(forn).strip(),
                tipo=None,
                numero=None,
                danfe=None,
                emissao=r[hmap.get('D. DOCUMENTO')] if 'D. DOCUMENTO' in hmap else None,
                valor=round(float(vtot), 2),
                lancamento=r[hmap.get('D. LANÇAMENTO')] if 'D. LANÇAMENTO' in hmap else (r[hmap.get('D. LANCAMENTO')] if 'D. LANCAMENTO' in hmap else None),
                portador=r[hmap.get('DESCRIÇÃO PORTADOR')] if 'DESCRIÇÃO PORTADOR' in hmap else (r[hmap.get('DESCRICAO PORTADOR')] if 'DESCRICAO PORTADOR' in hmap else None),
                origem=r[hmap.get('ORIGEM')] if 'ORIGEM' in hmap else None,
                imposto=is_tax,
            ))
    else:
        raise ValueError(f'Formato do Everest nao reconhecido. Cabecalhos: {list(hmap.keys())}')
    if exclude_taxes:
        rows = [r for r in rows if not r['imposto']]
    return rows, tipo_export

def load_fatura_itau(path):
    """Carrega fatura fechada Itau (formato .xlsx padrao Itaú)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    # achar linha do header 'Data | Lancamento | Parcelamento | Valor'
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [str(c).upper().strip() if c else '' for c in row]
        if 'DATA' in vals and 'LANÇAMENTO' in vals and 'VALOR' in vals:
            header_row = i
            break
    if header_row is None:
        # fallback: procurar 'LANCAMENTO'
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [str(c).upper().strip() if c else '' for c in row]
            if 'DATA' in vals and 'LANCAMENTO' in vals and 'VALOR' in vals:
                header_row = i
                break
    if header_row is None:
        raise ValueError(f'Nao achei header da fatura em {path}')
    # posicoes das colunas dentro do header row (pode ter Nones a esquerda)
    hdr_row = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    def col(name):
        for i, c in enumerate(hdr_row):
            if c and str(c).upper().strip() == name:
                return i
        return None
    ci_data = col('DATA')
    ci_lanc = col('LANÇAMENTO') or col('LANCAMENTO')
    ci_parc = col('PARCELAMENTO')
    ci_val = col('VALOR')
    ci_cart = col('NÚMERO DO CARTÃO') or col('NUMERO DO CARTAO')
    rows = []
    for r in ws.iter_rows(min_row=header_row+1, values_only=True):
        if ci_data is None or ci_lanc is None or ci_val is None: continue
        if r[ci_data] is None or r[ci_lanc] is None or r[ci_val] is None: continue
        data = r[ci_data]
        lanc = r[ci_lanc]
        parcel = r[ci_parc] if ci_parc is not None else None
        try:
            v = round(float(r[ci_val]), 2)
        except (TypeError, ValueError):
            continue
        cartao = r[ci_cart] if ci_cart is not None else None
        lower = str(lanc).lower()
        if 'pagamento efetuado' in lower: tipo = 'pagamento'
        elif 'anuidade' in lower: tipo = 'anuidade'
        elif any(k in lower for k in ('iof','multa por atraso','encargos de atraso','juros de')): tipo = 'encargo'
        else: tipo = 'transacao'
        rows.append(dict(data=data, lanc=str(lanc).strip(), parcel=parcel,
                          valor=v, cartao=cartao, tipo=tipo))
    return rows

# =============================================================================
# ENGINE DE MATCH
# =============================================================================
def to_date(x):
    if isinstance(x, datetime): return x.date()
    if isinstance(x, date): return x
    return None

def days_diff(a, b):
    da, db = to_date(a), to_date(b)
    if da is None or db is None: return 999
    return abs((da - db).days)

def match_two_lists(list_a, list_b, aliases_map,
                    key_a_valor='valor', key_b_valor='valor',
                    key_a_data='data', key_b_data='data',
                    key_a_forn='fornecedor', key_b_forn='fornecedor',
                    tolerance_abs=50.0, tolerance_pct=0.05, max_days=10,
                    min_supplier_score=0.55):
    """Faz o match multiset em 2 rounds:
      1) valor exato (multiset)
      2) mesmo fornecedor (score >= min_supplier_score OU alias canonico igual)
         + valor dentro de tolerancia + data <= max_days

    Retorna (matched, near_matches, sobra_a, sobra_b)
      matched:      lista de (a_row, b_row) com match exato
      near_matches: lista de (a_row, b_row, dv, dd, score) com match aproximado
      sobra_a:      itens de A sem match
      sobra_b:      itens de B sem match
    """
    # Round 1: valor exato multiset
    a_by = defaultdict(list); b_by = defaultdict(list)
    for i, x in enumerate(list_a): a_by[x[key_a_valor]].append(i)
    for j, y in enumerate(list_b): b_by[y[key_b_valor]].append(j)
    used_a = set(); used_b = set()
    matched = []
    for v in set(list(a_by) + list(b_by)):
        ai = a_by[v]; bj = b_by[v]
        for k in range(min(len(ai), len(bj))):
            matched.append((list_a[ai[k]], list_b[bj[k]]))
            used_a.add(ai[k]); used_b.add(bj[k])
    # Round 2: near-match por fornecedor
    remain_a = [(i, list_a[i]) for i in range(len(list_a)) if i not in used_a]
    remain_b = [(j, list_b[j]) for j in range(len(list_b)) if j not in used_b]
    near_matches = []
    used_b2 = set()
    # ordenar por valor descendente para pegar itens grandes primeiro
    remain_a.sort(key=lambda x: -x[1][key_a_valor])
    for i, a in remain_a:
        best = None  # (score_tuple, j_index)
        va = a[key_a_valor]
        for j, b in remain_b:
            if j in used_b2: continue
            vb = b[key_b_valor]
            dv = abs(va - vb)
            tol = max(tolerance_abs, va * tolerance_pct)
            if dv > tol: continue
            dd = days_diff(a[key_a_data], b[key_b_data])
            if dd > max_days: continue
            score = supplier_similarity(a.get(key_a_forn), b.get(key_b_forn), aliases_map)
            if score < min_supplier_score: continue
            # Preferir maior score, depois menor dv, depois menor dd
            key = (-score, dv, dd)
            if best is None or key < best[0]:
                best = (key, j, b, dv, dd, score)
        if best is not None:
            _, j, b, dv, dd, score = best
            used_b2.add(j); used_a.add(i)
            near_matches.append((a, b, round(dv, 2), dd, round(score, 2)))
    # Sobras
    sobra_a = [list_a[i] for i in range(len(list_a)) if i not in used_a]
    sobra_b = [list_b[j] for j in range(len(list_b)) if j not in used_b and j not in used_b2]
    return matched, near_matches, sobra_a, sobra_b

# =============================================================================
# CLASSIFICADORES
# =============================================================================
def classify_excel_row(row):
    n = str(row.get('fornecedor') or '').upper()
    if any(k in n for k in SERV_NOME_EXCEL): return 'Servico/assinatura (sem DANFE)'
    if row.get('setor') in SERV_SETOR: return 'Servico/assinatura (sem DANFE)'
    return 'Deveria ter NF-e no Everest'

def classify_fatura_row(row):
    e = str(row.get('lanc') or '').upper()
    for k in SERVICO_FATURA:
        if k in e: return 'Servico/app (sem DANFE)'
    for k in COMB_ESTAC:
        if k in e: return 'Combustivel/estac (cupom)'
    return 'Mercadoria - conferir NF-e'

# =============================================================================
# DUPLICIDADES NO EXCEL
# =============================================================================
def find_duplicates(excel):
    grp = defaultdict(list)
    for r in excel:
        d = to_date(r.get('data'))
        grp[(r['valor'], d)].append(r)
    dups = []
    for (val, d), rows in grp.items():
        if len(rows) < 2: continue
        forns = {str(x.get('fornecedor')).upper() for x in rows}
        tipo = 'Fornecedores diferentes (mesma compra em 2 nomes?)' if len(forns) >= 2 \
               else 'Mesmo fornecedor (lancamento em dobro?)'
        for x in rows:
            dd = dict(x); dd['tipo_dup'] = tipo
            dups.append(dd)
    dups.sort(key=lambda r: (-r['valor'], str(r.get('data'))))
    return dups

# =============================================================================
# ESCRITA DO XLSX
# =============================================================================
def build_xlsx(out_path, mes_label, excel, everest, fatura,
               recon_excel_everest, recon_fatura_excel, recon_fatura_everest,
               dups, aliases_used):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR = Font(bold=True, color='FFFFFF'); HFILL = PatternFill('solid', fgColor='2F5496')
    TITLE = Font(bold=True, size=13); BOLD = Font(bold=True)
    WARN = PatternFill('solid', fgColor='FCE4D6'); OK = PatternFill('solid', fgColor='E2EFDA')
    GREY = PatternFill('solid', fgColor='F2F2F2'); ALERT = PatternFill('solid', fgColor='F4B084')
    NEAR = PatternFill('solid', fgColor='FFF2CC')
    thin = Side(style='thin', color='BFBFBF')
    BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
    MONEY = '#,##0.00'; DATE = 'dd/mm/yyyy'

    def hdr(ws, cols, r=1):
        for c, t in enumerate(cols, 1):
            cell = ws.cell(r, c, t); cell.font = HDR; cell.fill = HFILL; cell.border = BORD
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def widths(ws, w):
        for i, x in enumerate(w, 1):
            ws.column_dimensions[get_column_letter(i)].width = x

    wb = openpyxl.Workbook()

    # ============ Resumo ============
    r = wb.active; r.title = 'Resumo'
    r['A1'] = f'Conciliacao mensal Pe de Manga - {mes_label}'
    r['A1'].font = TITLE
    r['A2'] = 'Hierarquia: Excel (primario) x Everest (secundario) + Fatura (complemento)'
    r['A2'].font = Font(italic=True, color='808080')
    row = [4]  # mutavel
    def line(lbl, val, money=False, bold=False, fill=None):
        a = r.cell(row[0], 1, lbl); b = r.cell(row[0], 3, val)
        if bold: a.font = BOLD; b.font = BOLD
        if money: b.number_format = MONEY
        if fill:
            for c in (1, 2, 3): r.cell(row[0], c).fill = fill
        row[0] += 1

    line('== BASES ==', '', bold=True)
    line('Excel controle - qtde', len(excel), bold=True)
    line('Excel controle - valor R$', round(sum(x['valor'] for x in excel), 2), money=True)
    line('Everest - qtde', len(everest), bold=True)
    line('Everest - valor R$', round(sum(x['valor'] for x in everest), 2), money=True)
    if fatura:
        line('Fatura compras (+) - qtde', sum(1 for x in fatura if x['tipo'] == 'transacao' and x['valor'] > 0))
        line('Fatura compras (+) - valor R$', round(sum(x['valor'] for x in fatura if x['tipo'] == 'transacao' and x['valor'] > 0), 2), money=True)
    row[0] += 1

    # Recon A: Excel x Everest
    m, nm, sa, sb = recon_excel_everest
    line('== RECON A: Excel (primario) x Everest (secundario) ==', '', bold=True)
    line('Casam valor exato', len(m), bold=True, fill=OK)
    line('Casam near (mesmo fornecedor, valor prox)', len(nm), bold=True, fill=NEAR)
    line('Excel SEM Everest (falta lancar no Everest)', len(sa), bold=True, fill=WARN)
    line('   Excel SEM Everest - valor R$', round(sum(x['valor'] for x in sa), 2), money=True, fill=WARN)
    line('Everest SEM Excel (falta lancar no Excel)', len(sb), bold=True, fill=WARN)
    line('   Everest SEM Excel - valor R$', round(sum(x['valor'] for x in sb), 2), money=True, fill=WARN)
    row[0] += 1

    if fatura:
        # Recon B: Fatura x Excel
        m, nm, sa, sb = recon_fatura_excel
        line('== RECON B: Fatura x Excel (primario) ==', '', bold=True)
        line('Casam valor exato', len(m), bold=True, fill=OK)
        line('Casam near', len(nm), bold=True, fill=NEAR)
        line('Fatura SEM Excel (falta lancar no Excel)', len(sa), bold=True, fill=WARN)
        line('   valor R$', round(sum(x['valor'] for x in sa), 2), money=True, fill=WARN)
        row[0] += 1

        # Recon C: Fatura x Everest
        m, nm, sa, sb = recon_fatura_everest
        line('== RECON C: Fatura x Everest (secundario) ==', '', bold=True)
        line('Casam valor exato', len(m), bold=True, fill=OK)
        line('Casam near', len(nm), bold=True, fill=NEAR)
        line('Fatura SEM Everest (falta lancar no Everest)', len(sa), bold=True, fill=WARN)
        line('   valor R$', round(sum(x['valor'] for x in sa), 2), money=True, fill=WARN)
        row[0] += 1

    line('Duplicidades no Excel', len(dups), bold=True, fill=WARN if dups else OK)
    row[0] += 2

    obs = [
        'OBSERVACOES:',
        '- Excel eh o CONTROLE PRIMARIO (source of truth). Todo lancamento operacional deve estar aqui.',
        '- Everest eh o CONTROLE SECUNDARIO (deveria bater com o Excel).',
        '- Fatura Itau eh COMPLEMENTO: tudo que aparece nela ja deveria estar em Excel e Everest.',
        '- "Near-match" (fundo amarelo) = mesmo fornecedor com valor levemente diferente. Confirmar manualmente.',
        f'- Aliases usados: {aliases_used} equivalencias conhecidas (data/supplier_aliases.json - editar para melhorar futuras conciliacoes).',
    ]
    for txt in obs:
        cell = r.cell(row[0], 1, txt)
        if txt.startswith('OBSERVACOES'): cell.font = BOLD
        else: cell.font = Font(italic=True, size=9, color='808080')
        row[0] += 1
    widths(r, [60, 8, 18])

    # ============ Sheets de detalhe ============
    def write_excel_sheet(ws_name, rows, extra_cols=None):
        ws = wb.create_sheet(ws_name)
        cols = ['Fornecedor (Excel)', 'Data', 'Valor R$', 'Setor', 'Classificacao']
        hdr(ws, cols)
        for i, x in enumerate(rows, 2):
            ws.cell(i, 1, x.get('fornecedor'))
            c = ws.cell(i, 2, x.get('data')); c.number_format = DATE
            c = ws.cell(i, 3, x['valor']); c.number_format = MONEY
            ws.cell(i, 4, x.get('setor'))
            cls = classify_excel_row(x)
            ws.cell(i, 5, cls)
            fill = ALERT if cls.startswith('Deveria') else None
            for cc in range(1, 6):
                ws.cell(i, cc).border = BORD
                if fill: ws.cell(i, cc).fill = fill
        ws.freeze_panes = 'A2'; widths(ws, [46, 12, 14, 18, 42])

    def write_everest_sheet(ws_name, rows):
        ws = wb.create_sheet(ws_name)
        cols = ['Fornecedor (Everest)', 'Emissao/Doc', 'Lancamento', 'Valor R$', 'Portador', 'Origem']
        hdr(ws, cols)
        for i, x in enumerate(rows, 2):
            ws.cell(i, 1, x.get('fornecedor'))
            c = ws.cell(i, 2, x.get('emissao')); c.number_format = DATE
            c = ws.cell(i, 3, x.get('lancamento')); c.number_format = DATE
            c = ws.cell(i, 4, x['valor']); c.number_format = MONEY
            ws.cell(i, 5, x.get('portador'))
            ws.cell(i, 6, x.get('origem'))
            for cc in range(1, 7): ws.cell(i, cc).border = BORD
            if x['valor'] >= 1000:
                for cc in range(1, 7): ws.cell(i, cc).fill = WARN
        ws.freeze_panes = 'A2'; widths(ws, [44, 14, 14, 14, 20, 12])

    def write_fatura_sheet(ws_name, rows):
        ws = wb.create_sheet(ws_name)
        cols = ['Data', 'Lancamento', 'Parcelamento', 'Valor R$', 'Cartao', 'Classificacao']
        hdr(ws, cols)
        for i, x in enumerate(rows, 2):
            c = ws.cell(i, 1, x.get('data')); c.number_format = DATE
            ws.cell(i, 2, x.get('lanc'))
            ws.cell(i, 3, x.get('parcel'))
            c = ws.cell(i, 4, x['valor']); c.number_format = MONEY
            ws.cell(i, 5, x.get('cartao'))
            cls = classify_fatura_row(x)
            ws.cell(i, 6, cls)
            fill = ALERT if 'Mercadoria' in cls else (WARN if 'Combust' in cls else GREY)
            for cc in range(1, 7): ws.cell(i, cc).border = BORD; ws.cell(i, cc).fill = fill
        ws.freeze_panes = 'A2'; widths(ws, [12, 36, 26, 14, 14, 42])

    def write_near_sheet(ws_name, near_matches, side_a='A', side_b='B',
                          key_a_forn='fornecedor', key_a_val='valor', key_a_date='data',
                          key_b_forn='fornecedor', key_b_val='valor', key_b_date='data'):
        ws = wb.create_sheet(ws_name)
        cols = [f'{side_a} Fornecedor', f'{side_a} Data', f'{side_a} Valor R$',
                f'{side_b} Fornecedor', f'{side_b} Data', f'{side_b} Valor R$',
                'Dif R$', 'Dif dias', 'Score fornec.']
        hdr(ws, cols)
        for i, (a, b, dv, dd, sc) in enumerate(near_matches, 2):
            ws.cell(i, 1, a.get(key_a_forn) if key_a_forn != 'lanc' else a.get('lanc'))
            c = ws.cell(i, 2, a.get(key_a_date)); c.number_format = DATE
            c = ws.cell(i, 3, a[key_a_val]); c.number_format = MONEY
            ws.cell(i, 4, b.get(key_b_forn) if key_b_forn != 'lanc' else b.get('lanc'))
            c = ws.cell(i, 5, b.get(key_b_date)); c.number_format = DATE
            c = ws.cell(i, 6, b[key_b_val]); c.number_format = MONEY
            c = ws.cell(i, 7, dv); c.number_format = MONEY
            ws.cell(i, 8, dd); ws.cell(i, 9, sc)
            for cc in range(1, 10): ws.cell(i, cc).border = BORD; ws.cell(i, cc).fill = NEAR
        ws.freeze_panes = 'A2'; widths(ws, [38, 12, 14, 38, 12, 14, 10, 9, 12])

    # Recon A
    m, nm, sa, sb = recon_excel_everest
    write_excel_sheet('A1_Excel_falta_Everest', sa)
    write_everest_sheet('A2_Everest_falta_Excel', sb)
    if nm:
        write_near_sheet('A3_NearMatch_Excel_Everest', nm,
                          side_a='Excel', side_b='Everest',
                          key_a_forn='fornecedor', key_a_val='valor', key_a_date='data',
                          key_b_forn='fornecedor', key_b_val='valor', key_b_date='emissao')

    if fatura:
        # Recon B: Fatura x Excel
        m, nm, sa, sb = recon_fatura_excel
        write_fatura_sheet('B1_Fatura_falta_Excel', sa)
        if nm:
            write_near_sheet('B2_NearMatch_Fatura_Excel', nm,
                              side_a='Fatura', side_b='Excel',
                              key_a_forn='lanc', key_a_val='valor', key_a_date='data',
                              key_b_forn='fornecedor', key_b_val='valor', key_b_date='data')
        # Recon C: Fatura x Everest
        m, nm, sa, sb = recon_fatura_everest
        write_fatura_sheet('C1_Fatura_falta_Everest', sa)
        if nm:
            write_near_sheet('C2_NearMatch_Fatura_Everest', nm,
                              side_a='Fatura', side_b='Everest',
                              key_a_forn='lanc', key_a_val='valor', key_a_date='data',
                              key_b_forn='fornecedor', key_b_val='valor', key_b_date='emissao')

    # Duplicidades
    if dups:
        ws = wb.create_sheet('Z_Duplicidades_Excel')
        hdr(ws, ['Fornecedor', 'Data', 'Valor R$', 'Setor', 'Tipo'])
        for i, x in enumerate(dups, 2):
            ws.cell(i, 1, x.get('fornecedor'))
            c = ws.cell(i, 2, x.get('data')); c.number_format = DATE
            c = ws.cell(i, 3, x['valor']); c.number_format = MONEY
            ws.cell(i, 4, x.get('setor'))
            ws.cell(i, 5, x['tipo_dup'])
            for cc in range(1, 6): ws.cell(i, cc).border = BORD; ws.cell(i, cc).fill = WARN
        ws.freeze_panes = 'A2'; widths(ws, [46, 12, 14, 18, 42])

    wb.save(out_path)

# =============================================================================
# CLI
# =============================================================================
def main():
    ap = argparse.ArgumentParser(description='Conciliacao mensal Pe de Manga')
    ap.add_argument('--excel', required=True, help='Excel controle (primario)')
    ap.add_argument('--everest', help='Everest export (Manutencao Notas ou Carteira Titulos)')
    ap.add_argument('--fatura', help='Fatura fechada Itau (xlsx)')
    ap.add_argument('--out', required=True, help='Caminho do xlsx de saida')
    ap.add_argument('--mes', default='', help='Rotulo do mes ex "07/2026"')
    ap.add_argument('--no-taxes', action='store_true', help='Excluir impostos do Everest (aplica se Carteira de Titulos)')
    ap.add_argument('--supplier-threshold', type=float, default=0.55, help='Score minimo de similaridade de fornecedor (0..1)')
    ap.add_argument('--tolerance-abs', type=float, default=50.0)
    ap.add_argument('--tolerance-pct', type=float, default=0.05)
    ap.add_argument('--max-days', type=int, default=10)
    args = ap.parse_args()

    aliases = load_aliases()
    print(f'Aliases carregados: {len(aliases)} variantes -> canonicos')

    excel = load_excel_controle(args.excel)
    print(f'Excel:   {len(excel)} lancamentos, R$ {sum(x["valor"] for x in excel):,.2f}')

    everest = []
    if args.everest:
        everest, tipo = load_everest_notas(args.everest, exclude_taxes=args.no_taxes)
        print(f'Everest ({tipo}): {len(everest)} titulos, R$ {sum(x["valor"] for x in everest):,.2f}')

    fatura = []
    if args.fatura:
        fatura = load_fatura_itau(args.fatura)
        trans = [x for x in fatura if x['tipo'] == 'transacao' and x['valor'] > 0]
        print(f'Fatura: {len(fatura)} linhas, {len(trans)} compras R$ {sum(x["valor"] for x in trans):,.2f}')

    common_kwargs = dict(aliases_map=aliases, tolerance_abs=args.tolerance_abs,
                         tolerance_pct=args.tolerance_pct, max_days=args.max_days,
                         min_supplier_score=args.supplier_threshold)

    # Recon A: Excel x Everest
    if everest:
        recon_A = match_two_lists(excel, everest,
                                    key_a_forn='fornecedor', key_a_valor='valor', key_a_data='data',
                                    key_b_forn='fornecedor', key_b_valor='valor', key_b_data='emissao',
                                    **common_kwargs)
    else:
        recon_A = ([], [], excel, [])

    # Recon B: Fatura x Excel
    fatura_trans = [x for x in fatura if x['tipo'] == 'transacao' and x['valor'] > 0]
    if fatura_trans and excel:
        recon_B = match_two_lists(fatura_trans, excel,
                                    key_a_forn='lanc', key_a_valor='valor', key_a_data='data',
                                    key_b_forn='fornecedor', key_b_valor='valor', key_b_data='data',
                                    **common_kwargs)
    else:
        recon_B = ([], [], fatura_trans, [])

    # Recon C: Fatura x Everest
    if fatura_trans and everest:
        recon_C = match_two_lists(fatura_trans, everest,
                                    key_a_forn='lanc', key_a_valor='valor', key_a_data='data',
                                    key_b_forn='fornecedor', key_b_valor='valor', key_b_data='emissao',
                                    **common_kwargs)
    else:
        recon_C = ([], [], fatura_trans, [])

    dups = find_duplicates(excel)

    print(f'\n=== RECON A (Excel x Everest) ===')
    print(f'  Casam exato: {len(recon_A[0])}  Near: {len(recon_A[1])}')
    print(f'  Excel SEM Everest: {len(recon_A[2])} - R$ {sum(x["valor"] for x in recon_A[2]):,.2f}')
    print(f'  Everest SEM Excel: {len(recon_A[3])} - R$ {sum(x["valor"] for x in recon_A[3]):,.2f}')
    if fatura_trans:
        print(f'=== RECON B (Fatura x Excel) ===')
        print(f'  Casam exato: {len(recon_B[0])}  Near: {len(recon_B[1])}')
        print(f'  Fatura SEM Excel: {len(recon_B[2])} - R$ {sum(x["valor"] for x in recon_B[2]):,.2f}')
        print(f'=== RECON C (Fatura x Everest) ===')
        print(f'  Casam exato: {len(recon_C[0])}  Near: {len(recon_C[1])}')
        print(f'  Fatura SEM Everest: {len(recon_C[2])} - R$ {sum(x["valor"] for x in recon_C[2]):,.2f}')
    print(f'Duplicidades Excel: {len(dups)}')

    mes = args.mes or datetime.today().strftime('%m/%Y')
    build_xlsx(args.out, mes, excel, everest, fatura,
                recon_A, recon_B, recon_C, dups, aliases_used=len(aliases))
    print(f'\nSalvo: {args.out}')

if __name__ == '__main__':
    main()
